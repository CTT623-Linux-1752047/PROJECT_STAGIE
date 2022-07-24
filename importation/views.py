from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from common.Const import Const
from common.Message import Message 
from common.Log import Log
from components.models.variable import Variable
from django.http import HttpResponse, JsonResponse


def index(request):
    if request.method == 'POST':
        try: 
            return JsonResponse({"response": request}, status=200)
            file =  request.FILES['myfile']
            valid = validateFileExcel(file)
            if(valid["flag"]):
                return JsonResponse({"response": True}, status=200)
            return JsonResponse({"response": False}, status=200)
        except Exception as e: 
            return JsonResponse({"response": e}, status=200)
        
    return render(request, 'importer/index.html')


# function validate excel is correct template
# return : object {flag(false if has error) , listMessageError} 
def validateFileExcel(file):
    return True
    wb = load_workbook(file, read_only=True)
    
    # case check file excel is not exist 8 sheet Variable Value, Variable, Technique, Task Type, Level, Group, Student, LocalInsInfo
    if (Const.SHEET_NAME_VARIABLE not in wb.sheetnames
        or Const.SHEET_NAME_VARIABLE_VALUE not in wb.sheetnames
        or Const.SHEET_NAME_TASK_TYPE not in wb.sheetnames
        or Const.SHEET_NAME_TECHNIQUE not in wb.sheetnames
        or Const.SHEET_NAME_LEVEL not in wb.sheetnames
        or Const.SHEET_NAME_GROUP not in wb.sheetnames
        or Const.SHEET_NAME_STUDENT not in wb.sheetnames
        or Const.SHEET_NAME_LOCALINSINFO not in wb.sheetnames):
            return {
                "flag" : False,
                "listMessage" : [Message.MSG_013]
                    }   
    # else:
    #     # case check in 8 sheet is exist column ID (column A)
    #     resultTechnique = checkColumnAIsCorrect(file, Const.SHEET_NAME_TECHNIQUE)
    #     resultVariableValue = checkColumnAIsCorrect(file, Const.SHEET_NAME_VARIABLE_VALUE)
    #     resultVariable = checkColumnAIsCorrect(file, Const.SHEET_NAME_VARIABLE)
    #     resultTaskType = checkColumnAIsCorrect(file, Const.SHEET_NAME_TASK_TYPE)
    #     resultLevel = checkColumnAIsCorrect(file, Const.SHEET_NAME_LEVEL)
    #     resultGroup = checkColumnAIsCorrect(file, Const.SHEET_NAME_GROUP)
    #     resultLocalInsInfo = checkColumnAIsCorrect(file, Const.SHEET_NAME_LOCALINSINFO)
    #     resultStudent = checkColumnAIsCorrect(file, Const.SHEET_NAME_STUDENT)
    #     if (resultTechnique["flag"] == True 
    #         and resultTaskType["flag"]  == True 
    #         and resultVariable["flag"] == True 
    #         and resultVariableValue["flag"] == True
    #         and resultStudent["flag"] == True 
    #         and resultGroup["flag"] == True 
    #         and resultLocalInsInfo["flag"] == True
    #         and resultLevel["flag"] == True):
    #         referenceTaskType = checkReferenceTaskType(file)
    #         referenceVariableValue = checkReferenceVariableValue(file)
    #         referenceTechnique = checkReferenceTechnique(file)
    #         referenceStudent = checkReferenceStudent(file)
    #         referenceLocalInsInfo = checkReferenceLocalInsInfo(file)
           
    #         if (referenceTechnique["flag"] == True and referenceTaskType["flag"] == True and referenceVariableValue["flag"] == True and referenceStudent["flag"] == True and referenceLocalInsInfo["flag"] == True) :
    #             return {
    #             "flag" : True,
    #             "listMessage" : []
    #             }   
    #         else :
    #             return {
    #             "flag" : False,
    #             "listMessage" : referenceTaskType["listMessage"] + 
    #             referenceVariableValue["listMessage"] + 
    #             referenceTechnique["listMessage"] + 
    #             referenceStudent["listMessage"] + 
    #             referenceLocalInsInfo["listMessage"]
    #             }   
    #     else:
    #         return {
    #             "flag" : False,
    #             "listMessage" : resultTechnique["listMessage"] + 
    #             resultVariable["listMessage"] + 
    #             resultVariableValue["listMessage"] + 
    #             resultStudent["listMessage"] + 
    #             resultGroup["listMessage"] + 
    #             resultLocalInsInfo["listMessage"] + 
    #             resultLevel["listMessage"] + 
    #             resultTaskType["listMessage"]
    #             }   
    return True

# function is check column A is correct
def checkColumnAIsCorrect(file, sheetname):
    try:
        workBook = load_workbook(file)
        workSheet = workBook[sheetname]
        row = 0
        flg = True
        msg = ""
        lstRow = []
        # Check column A is not empty cell in between 2 cell
        for cell in workSheet['A']:
            row = row + 1
            if (cell.value is None) :
                if (row + 1 < (len(workSheet['A']) - 1)):
                    if (workSheet['A'][row + 1] is not None and workSheet['A'][row + 1] != "") :
                        lstRow.append(str(row))
                        flg = False
                else:
                    lstRow.append(str(row))
                    flg = False
        workBook.close()
        if not flg :
            # write error 
            msg = Log.writeNotification(Message.MSG_001 , [sheetname, ",".join(lstRow)])
        return {
            "flag": flg,
            "listMessage": [msg]
            }
    except Exception as e:
        # write log error 
        Log.writeLog("Validate > checkColumnANotEmpty() " + (e))

# function is check reference of sheet Task Type
def checkReferenceTaskType(file):
    result = False
    listMessage = []
    try :
        wb = load_workbook(file)
        cellFirstSubTask = ""
        cellLastSubTask =""

        # Check number of column variabel's tasktype match with code of variable
        cntCol = 0
        checkIsCellSubTask = False
        ws = wb[Const.SHEET_NAME_VARIABLE]
        cntRow = countRowColumn(ws)

        ws = wb[Const.SHEET_NAME_TASK_TYPE]
        for col in ws[1]:
            if (col.value == Const.SUB_TASK) :
                cellFirstSubTask = col.coordinate
                checkIsCellSubTask = True
            if (col.value is None and not checkIsCellSubTask) :
                cntCol = cntCol + 1
            cellLastSubTask = col.coordinate

        if cntCol == cntRow - 1 :
            result = True
        else:
            listMessage.append(Log.writeNotification(Const.MSG_002)) 

        # check code sub task type is exits in task type
        lstAllTaskType = getValuesColumn(wb[Const.SHEET_NAME_TASK_TYPE], 'A')
        lstAllTaskType = lstAllTaskType if lstAllTaskType != False else []
        cellFirstSubTask = cellFirstSubTask.replace("1","")
        cellLastSubTask = cellLastSubTask.replace("1","")
        lstSubTask = []

        for row in (ws[cellFirstSubTask + ":" + cellLastSubTask]):
            for cell in row :
                if not None and cell.value not in lstSubTask:
                    lstSubTask.append(cell.value)
        lstDiff = list(set(lstSubTask) - set(lstAllTaskType))
        lstDiff.remove(Const.SUB_TASK)
        lstDiff.remove(None)

        if (len(lstDiff) > 0):
            result = False
            listMessage.append(Log.writeNotification(Const.MSG_003, [",".join(lstDiff)]))
        wb.close()
    except Exception as e:
        Log.writeLog("Validate > checkReferenceVariableValue() " + (e))

    return {
        "flag" : result,
        "listMessage" : listMessage
        }

# function is check reference of sheet Variable Value
def checkReferenceVariableValue(file):
    result = False
    listMessage = []
    
    try :
        wb = load_workbook(file)

        #check code variable of variable value isn't match with code variable of sheet variable
        lstVariablesOfVariableSheet = getValuesColumn(wb[Const.SHEET_NAME_VARIABLE], 'A')
        lstVariableOfVariableValueSheet = getValuesColumn(wb[Const.SHEET_NAME_VARIABLE_VALUE], 'A')
        listDiff = list(set(lstVariablesOfVariableSheet) - set(lstVariableOfVariableValueSheet))
        listDiff = [i for i in listDiff if i]

        if (len(listDiff) == 0) :
            result = True
        else:
            listMessage.append(Log.writeNotification(Const.MSG_004, [",".join(listDiff)]))

        wb.close()
    except Exception as e:
        Log.writeLog("Validate > checkReferenceVariableValue() " + str(e))

    return {
        "flag" : result,
        "listMessage" : listMessage
        }

# function is check reference of sheet technique
def checkReferenceTechnique(file):
    result = False
    listMessage = []
    try :
        wb = load_workbook(file)

        # check code task type of sheet Task Type is exist in code of sheet Task Type
        lstAllTTofSheetTaskType = getValuesColumn(wb[Const.SHEET_NAME_TASK_TYPE], 'A')
        lstAllPPragOfSheetTechnique = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'C')
        lstAllPPragOfSheetTechnique = [i for i in lstAllPPragOfSheetTechnique if i]
        lstAllTTofSheetTaskType = [i for i in lstAllTTofSheetTaskType if i]

        lstDiff = list(set(lstAllTTofSheetTaskType if not lstAllTTofSheetTaskType else []) - set(lstAllPPragOfSheetTechnique if not lstAllPPragOfSheetTechnique else []))
        if len(lstDiff) == 0 :
            result = True
        else :
            result = False
            listMessage.append(Log.writeNotification(Const.MSG_005, {",".join(lstDiff)}))

        # check code technique of sub technique in sheet Technique
        lstAllCodeTechnique = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'A')
        lstAllCodeTechnique = [i for i in lstAllCodeTechnique if i] if lstAllCodeTechnique != False else []

        # check for concurency global
        lstConcuGl = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'D')
        lstConcuGl = [i for i in lstConcuGl if i] if lstConcuGl != False else []
        lstConcuGl = convertCellTechniqueSheet(lstConcuGl)
        lstDiffConCuGl = list(set(lstConcuGl) - set(lstAllCodeTechnique))
        
        if 'None' in lstDiffConCuGl and len(lstDiffConCuGl) == 1:
            result = True
        else:
            result = False
            listMessage.append(Log.writeNotification(Const.MSG_006, [",".join(lstDiffConCuGl)]))

        # check for concurency inc
        lstConcuInc = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'E')
        lstConcuInc = [i for i in lstConcuInc if i] if lstConcuInc != False else []
        lstConcuInc = convertCellTechniqueSheet(lstConcuInc)
        lstDiffConCuInc = list(set(lstConcuInc) - set(lstAllCodeTechnique))
        
        if 'None' in  lstDiffConCuInc and len(lstDiffConCuInc) == 1:
            result = True
        else:
            result = False
            listMessage.append(Log.writeNotification(Const.MSG_007, [",".join(lstDiffConCuInc)]))

        # check for concurency ext
        lstConcuExt = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'F')
        lstConcuExt = [i for i in lstConcuExt if i] if lstConcuExt != False else []
        lstConcuExt = convertCellTechniqueSheet(lstConcuExt)
        lstDiffConCuExt = list(set(lstConcuExt) - set(lstAllCodeTechnique))
        
        if 'None' in lstDiffConCuExt and len(lstDiffConCuExt) == 1 :
            result = True
        else:
            result = False
            listMessage.append(Log.writeNotification(Const.MSG_008, [",".join(lstDiffConCuExt)]))

        wb.close()
    except Exception as e:
        Log.writeLog("Validate > checkReferenceTechnique() " + str(e))

    return {
    "flag" : result,
    "listMessage" : listMessage
    }

#  function is check reference of sheet student
def checkReferenceStudent(file):
    result = True
    listMessage = []
    try :
        wb = load_workbook(file)
        tmp = []
        # check code of group in sheet student has exist code in sheet Group
        lstCdGroupSheetStudent = getValuesColumn(wb[Const.SHEET_NAME_STUDENT], 'C')
        lstCdGroupSheetGroup = getValuesColumn(wb[Const.SHEET_NAME_GROUP], 'A')

        for i in lstCdGroupSheetStudent:
            if i not in lstCdGroupSheetGroup :
                tmp.append(i)
                result = False
        wb.close()
        if len(tmp) > 0 :
            listMessage.append(Log.writeNotification(Const.MSG_009, {",".join(tmp)}))
        return {
        "flag" : result,
        "listMessage" : listMessage
        }
    except Exception as e:
        listMessage.append(Log.writeLog("Validate > checkReferenceStudent() " + str(e)))

# function is check reference of sheet LocalInsInfo
def checkReferenceLocalInsInfo(file):
    result = True
    listMessage = []
    try:
        wb = load_workbook(file)
        lstCdStudent = []
        # check has code student is exist in sheet Student
        lstCdStudentSheetStudent = getValuesColumn(wb[Const.SHEET_NAME_STUDENT], 'A')
        lstCdStudentSheetLocalInsInfo = getValuesColumn(wb[Const.SHEET_NAME_LOCALINSINFO], 'B')
        i = 2
        for cdStudent in lstCdStudentSheetLocalInsInfo :
            if cdStudent not in lstCdStudentSheetStudent :
                result = False
                lstCdStudent.append("B" + str(i))
            i = i + 1
        if len(lstCdStudent) > 0 :
            listMessage.append(Log.writeNotification(Const.MSG_010, {",".join(lstCdStudent)}))

        # check has code level is exist in sheet Level
        lstLevelSheetLevel = getValuesColumn(wb[Const.SHEET_NAME_LEVEL], 'B')
        lstLevelSheetLocalInsInfo = getValuesColumn(wb[Const.SHEET_NAME_LOCALINSINFO], 'D')
        lstLevel = []
        i = 2
        for level in lstLevelSheetLocalInsInfo :
            if level.lower() not in [x.lower() for x in lstLevelSheetLevel]:
                result = False
                lstLevel.append("D" + str(i))
            i = i + 1
        if len(lstLevel) > 0 :
            listMessage.append(Log.writeNotification(Const.MSG_011, {",".join(lstLevel)}))

        # check has code technique is exist in sheet technique
        lstCdTechiqueSheetTechnique = getValuesColumn(wb[Const.SHEET_NAME_TECHNIQUE], 'A')
        lstCdTechniqeSheetLocalInsInfo = getValuesColumn(wb[Const.SHEET_NAME_LOCALINSINFO], 'C')
        i = 2
        lstTech = []
        for techs in lstCdTechniqeSheetLocalInsInfo:
            if len(techs.split(',')) > 1 :
                lstCdTech = techs.replace(" ","").split(',')
                for tech in lstCdTech :
                    if tech not in lstCdTechiqueSheetTechnique :
                        result = False
                        lstTech.append("C" + str(i))
            else :
                if techs not in lstCdTechiqueSheetTechnique:
                    result = False
                    lstTech.append("C" + str(i))

            i = i + 1
        if len(lstTech) > 0 :
            listMessage.append(Log.writeNotification(Const.MSG_012, {",".join(lstTech)}))

        wb.close()
        return {
        "flag" : result,
        "listMessage" : listMessage
        }
    except Exception as e:
        Log.writeLog("Validate > checkReferenceLocalInsInfo() " + str(e))

# function count row of column
def countRowColumn(workSheet):
    cnt = 0;
    for row in workSheet:
        cnt = cnt + 1
    return cnt

 # function is get all values of a column
def getValuesColumn(ws, columnName):
    if columnName.strip() != "" :
        lst =[]
        for col in ws[columnName]:
            lst.append(str(col.value).strip())
        del lst[0]
        return lst
    else :
        return False

# function support convert a value cell of sheet Technique to list (ex : cell A1 has value "TE001,TE003,TE004")
def convertCellTechniqueSheet(lst = []):
    result = []
    if len(lst) > 0 :
        for cell in lst :
            if len(str(cell).split(",")) > 1 :
                for item in str(cell).split(",") :
                    if item.strip() != "" : result.append(item.strip())
            else :
                result.append(str(cell).strip())

    return result

# function return data after load excel behind format
# return data{'Node': {'Variable':[], 'VariableValue':[], 'TaskType': [], 'Technique': []}, 'Relationship' : []}
def readExcel(file):
    try:
        data = {}
        # create node
        # data['Node'] = {Const.NAME_VARIABLE : getDataSheetVariable(file),
        #                 Const.NAME_VARIABLE_VALUE : self.getDataSheetVariableValue(),
        #                 Const.NAME_TASK_TYPE : self.getDataSheetTaskType(),
        #                 Const.NAME_TECHNIQUE : self.getDataSheetTechnique(),
        #                 Const.NAME_LEVEL : self.getDataSheetLevel(),
        #                 Const.NAME_STUDENT : self.getDataSheetStudent(),
        #                 Const.NAME_GROUP : self.getDataSheetGroup(),
        #                 Const.NAME_LOCALINSINFO : self.getDataSheetLocalInsInfo()}

        # # create relationship
        # data['Relationship'] = self.getRelationshipSheetTaskType() + self.getRelationshipSheetTechnique() \
        #                         + self.getRelationshipSheetVariableValue() + self.getRelationshipSheetStudent() \
        #                         + self.getRelationshipSheetLocalInsInfo()
        return data
    except Exception as e :
        Log.writeLog("Validate > readExcel() " + str(e))

# function get data node of sheet Variable
def createNodeVariable(file):
    try :
        lstVariable = []
        wb = load_workbook(file)
        lstCdVariable = getValuesColumn(wb[Const.SHEET_NAME_VARIABLE], 'A')
        lstCdVariable = lstCdVariable if lstCdVariable != False else []
        lstValueVariable = getValuesColumn(wb[Const.SHEET_NAME_VARIABLE], 'B')
        lstValueVariable = lstValueVariable if lstValueVariable != False else []

        for i in range(len(lstCdVariable)):
            Variable(variableCd=lstCdVariable[i], variableDescription=lstValueVariable[i]).save()
            # variable = 
            # lstVariable.append(variable)
        wb.close()
        return lstVariable
    except Exception as e :
        Log.writeLog("ExcelData > getDataSheetVariable() " + str(e))